import requests
import re
import os
import pymysql
import pymongo

from bs4 import BeautifulSoup
from openpyxl import Workbook
from datetime import datetime


def main():
    # url = "http://fund.eastmoney.com/daogou"
    # 连接中，通过"pi数字"跳转到第几页
    # url = "https://fund.eastmoney.com/daogou/#dt4;ft;rs;sd;ed;pr;cp;rt;tp;rk;se;nx;sc3y;stdesc;pi1;pn20;zfdiy;shlist"
    for i in range(1, 4):
        url = "https://fund.eastmoney.com/daogou/#dt4;ft;rs;sd;ed;pr;cp;rt;tp;rk;se;nx;sc3y;stdesc;pi" \
              + str(i) + ";pn20;zfdiy;shlist"
        # response = requests.get(url)
        # print(response.url)
        html = requests.get(url).content
        soup = BeautifulSoup(html, 'lxml')
        fund_infos = soup.find_all('tr', id=re.compile('^tr'))
        # print(fund_names)
        # for i in range(0, len(fund_names)):
        #     # print(len(fund_names))
        #     # print(fund_names[0].a['href'])
        #     fund_link = fund_names[i].a['href']
        #     fund_name = fund_names[i].a.text()
        #     print(fund_link)
        #     i = i+1
        wb = Workbook()
        # ws = wb.active
        fund_sheet = wb.create_sheet('fund_sheet', 0)
        fund_sheet['A1'] = '基金链接'
        fund_sheet['B1'] = '基金名称'
        fund_sheet['C1'] = '基金代码'
        fund_sheet['D1'] = '基金类型'
        fund_sheet['E1'] = '净值'
        fund_sheet['F1'] = '日增长率'
        fund_sheet['G1'] = '日期'
        fund_sheet['H1'] = '近1周'
        fund_sheet['I1'] = '近1月'
        fund_sheet['J1'] = '近3月'
        fund_sheet['K1'] = '近6月'
        fund_sheet['L1'] = '今年来'
        fund_sheet['M1'] = '近1年'
        fund_sheet['N1'] = '近2年'
        fund_sheet['O1'] = '近3年'
        fund_sheet['P1'] = '手续费链接'
        fund_sheet['Q1'] = '手续费'
        fund_sheet['R1'] = '购买起点'

        for fund_info in fund_infos:
            # 基金链接
            fund_link = fund_info.a['href']
            # print(fund_link)
            # 基金名称
            fund_name = fund_info.a.string
            # print(fund_name)
            # 基金代码
            fund_code = fund_info.span.string
            # print(fund_code)
            # 基金类型
            fund_type = fund_info.find('td', class_='fname').next_sibling.string
            # print(fund_type)
            # 基金净值
            fund_net_worth = fund_info.find('span', class_='ping').string
            # print(fund_net_worth)
            # 基金日增长率
            fund_daily_growth_rate = fund_info.find('span', class_='zhang').string
            # print(fund_daily_growth_rate)
            # 日期
            fund_date = fund_info.find('span', class_='gray').string
            # print(fund_date)
            # 近1周增长率
            fund_one_week = fund_info.contents[4].string
            # print(fund_one_week)
            # 近1月增长率
            fund_one_month = fund_info.contents[5].string
            # 近3月增长率
            fund_three_month = fund_info.contents[6].string
            # 近6月增长率
            fund_six_month = fund_info.contents[7].string
            # 今年来增长率
            fund_this_year = fund_info.contents[8].string
            # 近1年增长率
            fund_one_year = fund_info.contents[9].string
            # 近2年增长率
            fund_two_year = fund_info.contents[10].string
            # 近3年增长率
            fund_three_year = fund_info.contents[11].string
            # 手续费
            fund_rate_link = fund_info.find('div', class_='l').a['href']
            fund_rate = fund_info.find('div', class_='l').a.string
            # 购买起点
            fund_min_purchase_money = fund_info.find('div', class_='r').string

            # def save_excel():

            # 循环写入Excel，把这个页面的20条数据都写入Excel
            # openpyxl的append方法可以直接写入一行数据，而不需要循环
            # for col in fund_sheet.iter_cols(min_col=2, max_col=len(fund_infos)):

            row = [fund_link, fund_name, fund_code, fund_type, fund_net_worth, fund_daily_growth_rate,
                   fund_date, fund_one_week, fund_one_month, fund_three_month, fund_six_month, fund_this_year,
                   fund_one_year, fund_two_year, fund_three_year, fund_rate_link, fund_rate, fund_min_purchase_money]
            fund_sheet.append(row)

            # fund_sheet['Ai'] = fund_link
            # fund_sheet['Bi'] = fund_name
            # fund_sheet['Ci'] = fund_code
            # fund_sheet['Di'] = fund_type
            # fund_sheet['Ei'] = fund_net_worth
            # fund_sheet['Fi'] = fund_daily_growth_rate
            # fund_sheet['Gi'] = fund_date
            # fund_sheet['Hi'] = fund_one_week
            # fund_sheet['Ii'] = fund_one_month
            # fund_sheet['Ji'] = fund_three_month
            # fund_sheet['Ki'] = fund_six_month
            # fund_sheet['Li'] = fund_this_year
            # fund_sheet['Mi'] = fund_one_year
            # fund_sheet['Ni'] = fund_two_year
            # fund_sheet['Oi'] = fund_three_year
            # fund_sheet['Pi'] = fund_rate_link
            # fund_sheet['Qi'] = fund_rate
            # fund_sheet['Ri'] = fund_min_purchase_money

            # 保存excel
            dt = datetime.now()
            dt_format = dt.strftime('%Y%m%d%H%M')
            path = os.getcwd()
            # print(path)
            wb.save(path + '/' + dt_format + '.xlsx')

            # def save_mysql():
            # 保存数据到本地mysql数据库
            # 连接数据库
            connection = pymysql.connect(
                host='127.0.0.1',
                port=3306,
                user='root',
                password='',
                database='test'
            )
            # 建立游标，获取会话指针
            cursor = connection.cursor()
            # 创建数据库语句，插入数据
            sql = """insert into fund(fund_link, fund_name, fund_code, fund_type, fund_net_worth, fund_daily_growth_rate,
                                 fund_date, fund_one_week, fund_one_month, fund_three_month, fund_six_month, fund_this_year,
                                 fund_one_year, fund_two_year, fund_three_year, fund_rate_link, fund_rate,
                                 fund_min_purchase_money)
                        values (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)"""
            # 执行sql语句
            cursor.execute(sql, (fund_link, fund_name, fund_code, fund_type, fund_net_worth, fund_daily_growth_rate,
                                 fund_date, fund_one_week, fund_one_month, fund_three_month, fund_six_month, fund_this_year,
                                 fund_one_year, fund_two_year, fund_three_year, fund_rate_link, fund_rate,
                                 fund_min_purchase_money))
            # 提交
            connection.commit()
            # 关闭
            connection.close()

            # 数据存到MongoDB中
            # 数据库无密码连接，连接到MongoDB服务
            # "mongodb+srv://<username>:<password>@<cluster-name>.mongodb.net/myFirstDatabase"
            mongo_client = pymongo.MongoClient('mongodb://192.168.22.40:27017')
            # print(mongo_client.server_info())    # 判断是否链接上数据库
            # 连接到test数据库
            db_name = mongo_client.get_database('test')
            # 连接集合（相当于表），若不存在，则创建，若存在，则直接连接
            collection_name = db_name['jinjin']
            # 把数据组装成字典，赋值给info(相当于document)，然后插入数据（把document插入到collection中）
            info = {
                "fund_link": fund_link,
                "fund_name": fund_name,
                "fund_code": fund_code,
                "fund_code": fund_type,
                "fund_net_worth": fund_net_worth,
                "fund_daily_growth_rate": fund_daily_growth_rate,
                "fund_date": fund_date,
                "fund_one_week": fund_one_week,
                "fund_one_month": fund_one_month,
                "fund_three_month": fund_three_month,
                "fund_six_month": fund_six_month,
                "fund_this_year": fund_this_year,
                "fund_one_year": fund_one_year,
                "fund_two_year": fund_two_year,
                "fund_two_year": fund_three_year,
                "fund_rate_link": fund_rate_link,
                "fund_rate": fund_rate,
                "fund_min_purchase_money": fund_min_purchase_money
            }
            collection_name.insert_one(info)


if __name__ == '__main__':
    main()
